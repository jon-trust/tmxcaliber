from collections.abc import Iterable
from typing import Any

from openpyxl import load_workbook

from .cache import get_cached_local_path_for
from .errors import FrameworkNotFoundError

scf_versions: dict[str, dict[str, str]] = {
    "2023.4": {
        "url": (
            "https://github.com/securecontrolsframework/securecontrolsframework/"
            "raw/d1428c74aa76a66d9e131e6a3e3d1e61af25bd3a/"
            "Secure%20Controls%20Framework%20(SCF)%20-%202023.4.xlsx"
        ),
        "sheet_name": "SCF 2023.4",
    },
    "2024.1.1": {
        "url": (
            "https://github.com/securecontrolsframework/securecontrolsframework/"
            "raw/b14c2058fca9bb0085980cbb077b0bd3a71a09ea/"
            "Secure%20Controls%20Framework%20(SCF)%20-%202024.1.1.xlsx"
        ),
        "sheet_name": "SCF 2024.1.1",
    },
    "2025.3.1": {
        "url": (
            "https://github.com/securecontrolsframework/securecontrolsframework/"
            "raw/3c65d5907bc9802a310e2523a36dee7bb2fdebbe/"
            "secure-controls-framework-scf-2025-3-1.xlsx"
        ),
        "sheet_name": "SCF 2025.3.1",
    },
    # Add more versions as needed
}


def get_latest_supported_scf() -> str:
    return sorted(get_supported_scf())[-1]


def get_supported_scf() -> Iterable[str]:
    return scf_versions.keys()


def get_scf_config(version: str) -> dict[str, str]:
    if version in scf_versions:
        return scf_versions[version]
    raise ValueError("Unsupported SCF version requested")


def get_valid_scf_controls(version: str) -> list[str]:
    scf_data = get_full_scf_data(version)
    return [
        scf_id.strip()
        for scf_id in (row.get("SCF #") for row in scf_data)
        if isinstance(scf_id, str) and scf_id.strip()
    ]


def get_full_scf_data(version: str) -> list[dict[Any, Any]]:
    scf_config = get_scf_config(version)
    local_scf = get_cached_local_path_for(scf_config["url"])
    workbook = load_workbook(local_scf, read_only=True, data_only=True)
    try:
        worksheet = workbook[scf_config["sheet_name"]]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            return []

        return [dict(zip(headers, row, strict=False)) for row in rows]
    finally:
        workbook.close()


def _normalize_header(value: object) -> object:
    if not isinstance(value, str):
        return value
    return value.replace("\n", " ").strip()


def _normalize_text(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    normalized = value.strip()
    return normalized or None


def get_scf_data(version: str, framework_name: str) -> list[tuple[str, str]]:
    scf_data = get_full_scf_data(version=version)
    normalized_headers: dict[object, object] = {}
    for row in scf_data:
        for header in row:
            normalized_headers.setdefault(_normalize_header(header), header)

    if framework_name not in normalized_headers:
        raise FrameworkNotFoundError(framework_name)

    framework_header = normalized_headers[framework_name]
    framework_rows: list[tuple[str, str]] = []
    for row in scf_data:
        scf_id = _normalize_text(row.get("SCF #"))
        framework_values = row.get(framework_header)
        if not scf_id or not isinstance(framework_values, str):
            continue

        for framework_value in framework_values.split("\n"):
            normalized_framework_value = _normalize_text(framework_value)
            if normalized_framework_value:
                framework_rows.append((scf_id, normalized_framework_value))

    return framework_rows
